import asyncio
import datetime
import json
import math
from io import BytesIO

import aiohttp
import aiofiles

from openpyxl.workbook import Workbook

from settings import BASE_DIR

API_URL = 'https://api.sograph.xyz/api/campaign/list'
DAILY_GEMS_URL = 'https://api.sograph.xyz/api/user/check/in'
PAGE_SIZE = 12


class SoQuest:
    def __init__(self, address: str, signature: str, loop=None):
        self.address = address
        self.signature = signature
        self.headers = {
            'address': self.address,
            'signature': self.signature
        }
        self.data = []
        self.upload_data = []
        self.loop = loop or asyncio.get_event_loop()

    async def parse_data(self) -> str | None:
        filename = None
        total_count = await self.__get_campaigns_count()
        if total_count > 0:
            total_pages = math.ceil(total_count / PAGE_SIZE)
            tasks = [self.__get_data_per_page(page) for page in range(1, total_pages + 1)]
            await asyncio.gather(*tasks)
            self.__process_data()
            filename = await self.__dump_xlsx()
        return filename

    async def collect_daily(self) -> str:
        async with aiohttp.ClientSession() as session:
            async with session.post(DAILY_GEMS_URL, headers=self.headers) as response:
                if response.status == 200:
                    data_string = await response.text()
                    data_json = json.loads(data_string)
                    message = data_json.get('message')
                    if message == 'Signed in today':
                        return '0'
                    if message == 'OK':
                        return '1'
                    if message == 'Please login':
                        return '2'
                    else:
                        return message
                else:
                    return '404'

    async def __get_campaigns_count(self) -> int:
        params = {
            'campaign_type': 'all',
            'reward_type': 'all',
            'status': 'active',
            'trending': '0',
            'verified': '0',
            'name': '',
            'page': str(1),
            'pagesize': str(PAGE_SIZE),
            'hide_completed': '1',
        }
        async with aiohttp.ClientSession() as session:
            async with session.get(API_URL, params=params, headers=self.headers) as response:
                if response.status == 200:
                    data = await response.json()
                    return int(data.get('data').get('total'))
                else:
                    return 0

    async def __get_data_per_page(self, page: int):
        params = {
            'campaign_type': 'all',
            'reward_type': 'all',
            'status': 'active',
            'trending': '0',
            'verified': '0',
            'name': '',
            'page': str(page),
            'pagesize': str(PAGE_SIZE),
            'hide_completed': '1',
        }
        async with aiohttp.ClientSession() as session:
            async with session.get(API_URL, params=params, headers=self.headers) as response:
                if response.status == 200:
                    data = await response.json()
                    self.data.extend(data.get('data').get('data'))

    def __process_data(self):
        for data in self.data:
            gems_count = 20 if data.get('is_verify') and data.get('is_recommend')\
                else (10 if data.get('is_verify')
                      else 1)

            end_timestamp = data.get('end_time')
            if end_timestamp:
                current_time = datetime.datetime.now()
                target_time = datetime.datetime.fromtimestamp(end_timestamp)
                time_difference = target_time - current_time
                hours_left = round(time_difference.total_seconds() / 3600, 2)
            else:
                hours_left = 'Нет ограничения'

            self.upload_data.append(
                {
                    'Кол-во гемов': gems_count,
                    'Ссылка': data.get('url'),
                    'Название кампании': data.get('space_name'),
                    'Кол-во заданий': data.get('task_count'),
                    'Тип призов': ', '.join(data.get("prize_types")),
                    'Осталось времени (ч.)': hours_left
                }
            )

    async def __dump_xlsx(self):
        cur_datetime = datetime.datetime.utcnow().strftime('%d_%m_%Y_%H_%M_%S')
        filename = BASE_DIR / 'assets' / f'result_{cur_datetime}.xlsx'

        wb = Workbook()

        gem_sheets = {}
        for row in self.upload_data:
            gem_count = row['Кол-во гемов']
            if gem_count not in gem_sheets:
                gem_sheets[gem_count] = wb.create_sheet(title=str(gem_count))
                sheet = gem_sheets[gem_count]
                sheet.append(list(row.keys()))
            sheet = gem_sheets[gem_count]
            sheet.append(list(row.values()))

        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        buffer = BytesIO()
        wb.save(buffer)

        async with aiofiles.open(filename, 'wb') as f:
            await f.write(buffer.getvalue())

        buffer.close()
        wb.close()
